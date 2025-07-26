import pandas as pd
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def read_excel(file_path):
    try:
        return pd.read_excel(file_path, sheet_name=None)
    except Exception as e:
        raise Exception(f"Erro ao ler planilha: {str(e)}")

def limpar_forma_vida_substrato(df):
    """Limpa e separa corretamente forma de vida e substrato, removendo duplicatas"""
    
    def clean_and_deduplicate(text):
        """Remove duplicatas de uma string separada por vírgulas"""
        if pd.isna(text) or not str(text).strip():
            return ""
        
        text = str(text)
        # Separar por vírgula e remover duplicatas
        items = []
        seen = set()
        for item in text.split(','):
            item = item.strip()
            item_lower = item.lower()
            # Filtrar títulos e duplicatas
            if (item and 
                item_lower not in seen and 
                item_lower not in ['forma de vida', 'substrato', 'nan'] and
                item != 'nan'):
                items.append(item)
                seen.add(item_lower)
        
        return ', '.join(items)
    
    if "Forma de Vida" in df.columns and "Substrato" in df.columns:
        for idx, row in df.iterrows():
            forma_vida = str(row["Forma de Vida"]) if pd.notna(row["Forma de Vida"]) else ""
            substrato = str(row["Substrato"]) if pd.notna(row["Substrato"]) else ""
            
            # Se forma de vida está vazia e substrato tem dados que parecem ser de forma de vida
            if not forma_vida.strip() and substrato.strip():
                # Separar dados que podem estar misturados
                linhas = [linha.strip() for linha in substrato.split('\n') if linha.strip()]
                
                forma_temp = []
                substrato_temp = []
                
                # Palavras-chave para identificar forma de vida
                palavras_forma = ['Arbusto', 'Árvore', 'Erva', 'Liana', 'Subarbusto', 'Trepadeira', 'Herbácea']
                # Palavras-chave para identificar substrato  
                palavras_substrato = ['Terrícola', 'Rupícola', 'Epífita', 'Aquática', 'Epifítica']
                
                for linha in linhas:
                    # Pular linhas que são títulos
                    if linha in ['Forma de Vida', 'Substrato']:
                        continue
                        
                    # Verificar se a linha contém palavras de forma de vida
                    if any(palavra in linha for palavra in palavras_forma):
                        forma_temp.append(linha)
                    # Verificar se a linha contém palavras de substrato
                    elif any(palavra in linha for palavra in palavras_substrato):
                        substrato_temp.append(linha)
                    else:
                        # Se não conseguir identificar, tentar pela posição
                        # Geralmente forma de vida vem primeiro
                        if not forma_temp and linha:
                            forma_temp.append(linha)
                        elif linha:
                            substrato_temp.append(linha)
                
                # Atualizar os valores removendo duplicatas
                if forma_temp:
                    df.at[idx, "Forma de Vida"] = clean_and_deduplicate(', '.join(forma_temp))
                if substrato_temp:
                    df.at[idx, "Substrato"] = clean_and_deduplicate(', '.join(substrato_temp))
                elif not substrato_temp and forma_temp:
                    # Se só temos dados de forma de vida, limpar substrato
                    df.at[idx, "Substrato"] = ""
            else:
                # Limpar duplicatas mesmo quando os dados estão nas colunas corretas
                df.at[idx, "Forma de Vida"] = clean_and_deduplicate(forma_vida)
                df.at[idx, "Substrato"] = clean_and_deduplicate(substrato)
    
    return df

# Modifique a função salvar_planilha para incluir a limpeza:

def salvar_planilha(dataframes_dict, colunas_personalizadas=None):
    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx *.xlsm")],
        title="Salvar planilha"
    )
    if not output_path:
        return

    # Mapeamento completo das colunas internas para nomes de exibição
    col_map = {
        "Nº": "Nº",
        "Nome Científico": "Nome Científico",
        "familia": "Família",
        "autor": "Autor",
        "reflora_link": "Link Reflora",
        "Status Nome": "Status Nome",
        "inconsistencia": "Inconsistências",
        "distribuicao_geografica": "Distribuição",
        "dominios_fitogeograficos": "Domínios Fitogeográficos",
        "tipos_vegetacao": "Tipos de Vegetação",
        "forma_vida": "Forma de Vida",
        "substrato": "Substrato",
        "origem": "Origem",     # Vai garantir que a coluna de origem seja mapeada
        "endemismo": "Endemismo"   # endemismo tambem
    }

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            # Renomear colunas conforme mapeamento
            df = df.rename(columns=col_map)
            
            # Ordem das colunas (ajustada para incluir Origem e Endemismo)
            col_order = [
                "Nº", 
                "Nome Científico", 
                "Família", 
                "Autor",
                "Link Reflora", 
                "Status Nome", 
                "Inconsistências",
                "Distribuição",           
                "Domínios Fitogeográficos",  
                "Tipos de Vegetação",     
                "Forma de Vida",          
                "Substrato",
                "Origem",
                "Endemismo"
            ]

            # Filtrar apenas colunas existentes
            col_order = [col for col in col_order if col in df.columns]
            
            # Reordenar colunas
            df = df[col_order]
            
            # Limpeza específica para Forma de Vida e Substrato
            df = limpar_forma_vida_substrato(df)
            
            # Limpeza adicional
            for col in ["Substrato", "Forma de Vida"]:
                if col in df.columns:
                    df[col] = df[col].astype(str)
                    df[col] = df[col].str.replace("Forma de Vida", "", regex=False)
                    df[col] = df[col].str.replace("Substrato", "", regex=False)
                    df[col] = df[col].str.strip()
                    df[col] = df[col].str.replace(r'^[,\s]*$', '', regex=True)
            
            # Salvar no Excel
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            
            # Formatação
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Formatar Status Nome
            if "Status Nome" in df.columns:
                status_col = df.columns.get_loc("Status Nome") + 1
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=status_col)
                    if cell.value == "Nome desatualizado":
                        cell.font = Font(color="FF0000", bold=True)
                    elif cell.value == "Nome válido":
                        cell.font = Font(color="008000", bold=True)
                    elif cell.value == "Possível sinônimo":
                        cell.font = Font(color="FF8C00", bold=True)
                    elif cell.value == "Nome inválido":
                        cell.font = Font(color="8B0000", bold=True)
            
            # Formatar Links
            if "Link Reflora" in df.columns:
                link_col = df.columns.get_loc("Link Reflora") + 1
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=link_col)
                    if "http" in str(cell.value):
                        cell.style = "Hyperlink"
                        cell.font = Font(underline="single", color="0563C1")
            
            # Formatar Inconsistências
            if "Inconsistências" in df.columns:
                inconsist_col = df.columns.get_loc("Inconsistências") + 1
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=inconsist_col)
                    if cell.value and str(cell.value).strip():
                        cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            
            # Ajustar largura das colunas
            for col in worksheet.columns:
                max_len = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_len:
                            max_len = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max(max_len + 2, 12), 80)
                worksheet.column_dimensions[column_letter].width = adjusted_width

    messagebox.showinfo("Sucesso", f"Planilha salva em:\n{output_path}")